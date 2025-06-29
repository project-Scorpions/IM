from PyQt5.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, 
                            QTableWidget, QTableWidgetItem, QComboBox, QLineEdit, 
                            QDateEdit, QSpinBox, QDoubleSpinBox, QFormLayout, 
                            QGroupBox, QTabWidget, QMessageBox, QHeaderView, QFrame,
                            QDialog, QDialogButtonBox, QTreeWidget, QTreeWidgetItem, QCheckBox)
from PyQt5.QtCore import Qt, QDate
from PyQt5.QtGui import QIcon, QFont, QColor

from database.db_connector import DatabaseConnection
from utils.auth import Authentication

class ProductDialog(QDialog):
    """Dialog for adding or editing products"""
    
    def __init__(self, parent=None, product_id=None):
        super().__init__(parent)
        self.db = DatabaseConnection()
        self.auth = Authentication()
        self.product_id = product_id
        self.user = parent.user
        self.init_ui()
        
        if product_id:
            self.setWindowTitle("Edit Product")
            self.load_product_data()
        else:
            self.setWindowTitle("Add New Product")
    
    def init_ui(self):
        """Initialize the UI components"""
        self.setMinimumWidth(500)
        
        layout = QVBoxLayout(self)
        
        # Form layout for product details
        form_layout = QFormLayout()
        
        # Product name
        self.name_input = QLineEdit()
        form_layout.addRow("Product Name*:", self.name_input)
        
        # Medication Type (Branded/Generic)
        self.type_combo = QComboBox()
        self.type_combo.addItem("Branded", False)
        self.type_combo.addItem("Generic", True)
        form_layout.addRow("Medication Type*:", self.type_combo)
        
        # Category
        self.category_combo = QComboBox()
        self.load_categories()
        form_layout.addRow("Category:", self.category_combo)
        
        # Unit Measurement
        self.unit_input = QLineEdit()
        self.unit_input.setPlaceholderText("e.g., mg, ml, g, tablet")
        form_layout.addRow("Unit Measurement:", self.unit_input)
        
        # Description
        self.description_input = QLineEdit()
        form_layout.addRow("Description:", self.description_input)
        
        # Unit price
        self.unit_price_input = QDoubleSpinBox()
        self.unit_price_input.setRange(0, 99999.99)
        self.unit_price_input.setDecimals(2)
        self.unit_price_input.setSingleStep(0.1)
        form_layout.addRow("Unit Price*:", self.unit_price_input)
        
        # Cost price
        self.cost_price_input = QDoubleSpinBox()
        self.cost_price_input.setRange(0, 99999.99)
        self.cost_price_input.setDecimals(2)
        self.cost_price_input.setSingleStep(0.1)
        form_layout.addRow("Cost Price*:", self.cost_price_input)
        
        # Stock quantity
        self.stock_input = QSpinBox()
        self.stock_input.setRange(0, 999999)
        form_layout.addRow("Stock Quantity*:", self.stock_input)
        
        # Expiry date
        self.expiry_input = QDateEdit()
        self.expiry_input.setCalendarPopup(True)
        self.expiry_input.setDate(QDate.currentDate().addMonths(12))  # Default to 1 year
        form_layout.addRow("Expiry Date:", self.expiry_input)
        
        # Reorder level
        self.reorder_input = QSpinBox()
        self.reorder_input.setRange(0, 9999)
        self.reorder_input.setValue(10)  # Default reorder level
        form_layout.addRow("Reorder Level:", self.reorder_input)
        
        # Supplier
        self.supplier_combo = QComboBox()
        self.load_suppliers()
        form_layout.addRow("Supplier:", self.supplier_combo)
        
        # Active status checkbox
        self.active_checkbox = QCheckBox("Product Active")
        self.active_checkbox.setChecked(True)  # Default to active
        form_layout.addRow("Status:", self.active_checkbox)
        
        # Add form layout to main layout
        layout.addLayout(form_layout)
        
        # Add note about required fields
        required_note = QLabel("* Required fields")
        required_note.setStyleSheet("color: red;")
        layout.addWidget(required_note)
        
        # Button box
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)
    
    def load_categories(self):
        """Load categories into combo box"""
        try:
            query = "SELECT category_id, name FROM categories ORDER BY name"
            categories = self.db.execute_query(query, fetchall=True)
            
            self.category_combo.clear()
            self.category_combo.addItem("-- Select Category --", None)
            
            for category_id, name in categories:
                self.category_combo.addItem(name, category_id)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load categories: {str(e)}")
    
    def load_suppliers(self):
        """Load suppliers into combo box"""
        try:
            query = "SELECT supplier_id, name FROM suppliers ORDER BY name"
            suppliers = self.db.execute_query(query, fetchall=True)
            
            self.supplier_combo.clear()
            self.supplier_combo.addItem("-- Select Supplier (Optional) --", None)
            
            for supplier_id, name in suppliers:
                self.supplier_combo.addItem(name, supplier_id)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load suppliers: {str(e)}")
    
    def load_product_data(self):
        """Load product data if editing existing product"""
        try:
            # Check if the database has the medication fields
            connection = self.db.get_connection()
            cursor = connection.cursor()
            
            try:
                cursor.execute("""
                    SELECT column_name 
                    FROM information_schema.columns 
                    WHERE table_name = 'products' AND column_name = 'is_generic'
                """)
                has_med_fields = cursor.fetchone() is not None
            except:
                has_med_fields = False
            finally:
                cursor.close()
                self.db.release_connection(connection)
                
            if has_med_fields:
                # Query with medication fields
                query = """
                    SELECT p.product_name, p.category_id, p.description, p.unit_price, 
                           p.cost_price, p.stock_quantity, p.expiry_date, p.reorder_level, 
                           p.supplier_id, p.is_active, p.is_generic, p.unit_measurement
                    FROM products p
                    WHERE p.product_id = %s
                """
            else:
                # Query without medication fields
                query = """
                    SELECT p.product_name, p.category_id, p.description, p.unit_price, 
                           p.cost_price, p.stock_quantity, p.expiry_date, p.reorder_level, 
                           p.supplier_id, p.is_active
                    FROM products p
                    WHERE p.product_id = %s
                """
            
            product = self.db.execute_query(query, (self.product_id,), fetchone=True)
            
            if product:
                self.name_input.setText(product[0])
                
                # Set category
                if product[1]:
                    index = self.category_combo.findData(product[1])
                    if index >= 0:
                        self.category_combo.setCurrentIndex(index)
            
                self.description_input.setText(product[2] or "")
                self.unit_price_input.setValue(float(product[3]))
                self.cost_price_input.setValue(float(product[4]))
                self.stock_input.setValue(int(product[5]))
                
                # Set expiry date if available
                if product[6]:
                    self.expiry_input.setDate(product[6])
                
                self.reorder_input.setValue(int(product[7]))
                
                # Set supplier if available
                if product[8]:
                    index = self.supplier_combo.findData(product[8])
                    if index >= 0:
                        self.supplier_combo.setCurrentIndex(index)
                
                # Set active status
                self.active_checkbox.setChecked(product[9] if product[9] is not None else True)
                
                # Set medication details if available
                if has_med_fields and len(product) > 10:
                    is_generic = product[10]
                    self.type_combo.setCurrentIndex(1 if is_generic else 0)
                    
                    unit_measurement = product[11]
                    self.unit_input.setText(unit_measurement or "")
                
            else:
                QMessageBox.warning(self, "Warning", "Product not found.")
                self.reject()
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load product data: {str(e)}")
            import traceback
            print(traceback.format_exc())
            self.reject()
    
    def accept(self):
        """Handle dialog acceptance (OK button)"""
        # Validate input
        product_name = self.name_input.text().strip()
        if not product_name:
            QMessageBox.warning(self, "Validation Error", "Product name is required.")
            return
        
        if self.unit_price_input.value() <= 0:
            QMessageBox.warning(self, "Validation Error", "Unit price must be greater than zero.")
            return
        
        if self.cost_price_input.value() <= 0:
            QMessageBox.warning(self, "Validation Error", "Cost price must be greater than zero.")
            return
        
        # Create direct database connection for this operation
        connection = None
        cursor = None
        try:
            # Get values
            category_id = self.category_combo.currentData()
            description = self.description_input.text().strip()
            unit_price = self.unit_price_input.value()
            cost_price = self.cost_price_input.value()
            stock_quantity = self.stock_input.value()
            expiry_date = self.expiry_input.date().toString("yyyy-MM-dd")
            reorder_level = self.reorder_input.value()
            supplier_id = self.supplier_combo.currentData()
            is_active = self.active_checkbox.isChecked()
            
            # Get medication type and unit measurement
            is_generic = self.type_combo.currentData()
            unit_measurement = self.unit_input.text().strip()
            
            # Get a direct connection
            connection = self.db.get_connection()
            cursor = connection.cursor()
            
            # Check if the database has the medication fields
            cursor.execute("""
                SELECT column_name 
                FROM information_schema.columns 
                WHERE table_name = 'products' AND column_name = 'is_generic'
            """)
            has_med_fields = cursor.fetchone() is not None
            
            # If fields don't exist, try to add them
            if not has_med_fields:
                try:
                    cursor.execute("""
                        ALTER TABLE products 
                        ADD COLUMN IF NOT EXISTS is_generic BOOLEAN DEFAULT FALSE,
                        ADD COLUMN IF NOT EXISTS unit_measurement VARCHAR(50)
                    """)
                    connection.commit()
                    has_med_fields = True
                except:
                    # If we can't alter the table, continue without the fields
                    connection.rollback()
                    has_med_fields = False
            
            if self.product_id:
                # Update existing product
                if has_med_fields:
                    # Update with medication fields
                    query = """
                        UPDATE products
                        SET product_name = %s, category_id = %s, description = %s,
                            unit_price = %s, cost_price = %s, stock_quantity = %s,
                            expiry_date = %s, reorder_level = %s, supplier_id = %s,
                            is_active = %s, updated_at = CURRENT_TIMESTAMP,
                            is_generic = %s, unit_measurement = %s
                        WHERE product_id = %s
                    """
                    cursor.execute(
                        query, 
                        (product_name, category_id, description, unit_price, cost_price, 
                         stock_quantity, expiry_date, reorder_level, supplier_id, is_active,
                         is_generic, unit_measurement, self.product_id)
                    )
                else:
                    # Update without medication fields
                    query = """
                        UPDATE products
                        SET product_name = %s, category_id = %s, description = %s,
                            unit_price = %s, cost_price = %s, stock_quantity = %s,
                            expiry_date = %s, reorder_level = %s, supplier_id = %s,
                            is_active = %s, updated_at = CURRENT_TIMESTAMP
                        WHERE product_id = %s
                    """
                    cursor.execute(
                        query, 
                        (product_name, category_id, description, unit_price, cost_price, 
                         stock_quantity, expiry_date, reorder_level, supplier_id, is_active,
                         self.product_id)
                    )
                
                # Log activity
                action_details = f"Updated product: {product_name}"
                if not is_active:
                    action_details += " (Deactivated)"
                    
                self.auth.log_activity(
                    self.user['user_id'],
                    "update",
                    "products",
                    self.product_id,
                    action_details
                )
                
            else:
                # Insert new product
                if has_med_fields:
                    # Insert with medication fields
                    query = """
                        INSERT INTO products
                        (product_name, category_id, description, unit_price, cost_price,
                         stock_quantity, expiry_date, reorder_level, supplier_id, is_active,
                         is_generic, unit_measurement)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        RETURNING product_id
                    """
                    cursor.execute(
                        query, 
                        (product_name, category_id, description, unit_price, cost_price, 
                         stock_quantity, expiry_date, reorder_level, supplier_id, is_active,
                         is_generic, unit_measurement)
                    )
                else:
                    # Insert without medication fields
                    query = """
                        INSERT INTO products
                        (product_name, category_id, description, unit_price, cost_price,
                         stock_quantity, expiry_date, reorder_level, supplier_id, is_active)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        RETURNING product_id
                    """
                    cursor.execute(
                        query, 
                        (product_name, category_id, description, unit_price, cost_price, 
                         stock_quantity, expiry_date, reorder_level, supplier_id, is_active)
                    )
                
                result = cursor.fetchone()
                new_product_id = result[0]
                
                # Log activity
                self.auth.log_activity(
                    self.user['user_id'],
                    "insert",
                    "products",
                    new_product_id,
                    f"Created new product: {product_name}"
                )
                
                # Print debug info
                print(f"Created new product with ID {new_product_id}: {product_name}")
            
            # Explicitly commit the transaction
            connection.commit()
            
            QMessageBox.information(
                self, 
                "Success", 
                f"Product '{product_name}' {'updated' if self.product_id else 'created'} successfully."
            )
            
            super().accept()
            
        except Exception as e:
            if connection:
                connection.rollback()
            print(f"Error in product accept method: {str(e)}")
            import traceback
            print(traceback.format_exc())
            QMessageBox.critical(self, "Error", f"Failed to save product: {str(e)}")
        finally:
            if cursor:
                cursor.close()
            if connection:
                self.db.release_connection(connection)


class CategoryDialog(QDialog):
    """Dialog for adding or editing categories"""
    
    def __init__(self, parent=None, category_id=None):
        super().__init__(parent)
        self.db = DatabaseConnection()
        self.auth = Authentication()
        self.category_id = category_id
        self.user = parent.user if hasattr(parent, 'user') else {'user_id': 1}  # Default to admin if no user
        self.init_ui()
        
        if category_id:
            self.setWindowTitle("Edit Category")
            self.load_category_data()
        else:
            self.setWindowTitle("Add New Category")
    
    def init_ui(self):
        """Initialize the UI components"""
        self.setMinimumWidth(400)
        
        layout = QVBoxLayout(self)
        
        # Form layout for category details
        form_layout = QFormLayout()
        
        # Category name
        self.name_input = QLineEdit()
        form_layout.addRow("Category Name*:", self.name_input)
        
        # Description
        self.description_input = QLineEdit()
        form_layout.addRow("Description:", self.description_input)
        
        # Add form layout to main layout
        layout.addLayout(form_layout)
        
        # Add note about required fields
        required_note = QLabel("* Required fields")
        required_note.setStyleSheet("color: red;")
        layout.addWidget(required_note)
        
        # Button box
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)
    
    def load_category_data(self):
        """Load category data if editing existing category"""
        try:
            query = "SELECT name, description FROM categories WHERE category_id = %s"
            category = self.db.execute_query(query, (self.category_id,), fetchone=True)
            
            if category:
                self.name_input.setText(category[0])
                self.description_input.setText(category[1] or "")
            else:
                QMessageBox.warning(self, "Warning", "Category not found.")
                self.reject()
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load category data: {str(e)}")
            self.reject()
    
    def accept(self):
        """Handle dialog acceptance (OK button)"""
        # Validate input
        category_name = self.name_input.text().strip()
        if not category_name:
            QMessageBox.warning(self, "Validation Error", "Category name is required.")
            return
        
        # Create direct database connection for this operation
        connection = None
        cursor = None
        try:
            # Get values
            description = self.description_input.text().strip()
            
            # Get a direct connection
            connection = self.db.get_connection()
            cursor = connection.cursor()
            
            if self.category_id:
                # Update existing category
                query = """
                    UPDATE categories
                    SET name = %s, description = %s, updated_at = CURRENT_TIMESTAMP
                    WHERE category_id = %s
                """
                cursor.execute(query, (category_name, description, self.category_id))
            else:
                # Insert new category
                query = """
                    INSERT INTO categories (name, description)
                    VALUES (%s, %s)
                    RETURNING category_id
                """
                cursor.execute(query, (category_name, description))
                result = cursor.fetchone()
                new_category_id = result[0]
                self.category_id = new_category_id  # Store for activity logging
            
            # Explicitly commit the transaction
            connection.commit()
            
            # Log activity after successful commit
            try:
                if self.category_id:
                    action_type = "update" if self.category_id else "insert"
                    self.auth.log_activity(
                        self.user['user_id'],
                        action_type,
                        "categories",
                        self.category_id,
                        f"{'Updated' if action_type == 'update' else 'Created new'} category: {category_name}"
                    )
            except Exception as log_error:
                # Just log the error but don't stop the process
                print(f"Activity logging error: {log_error}")
            
            QMessageBox.information(
                self, 
                "Success", 
                f"Category '{category_name}' {'updated' if self.category_id else 'created'} successfully."
            )
            
            super().accept()
            
        except Exception as e:
            if connection:
                connection.rollback()
            print(f"Error in accept method: {str(e)}")
            import traceback
            print(traceback.format_exc())
            QMessageBox.critical(self, "Error", f"Failed to save category: {str(e)}")
        finally:
            if cursor:
                cursor.close()
            if connection:
                self.db.release_connection(connection)


class InventoryManagementWidget(QWidget):
    """Widget for managing inventory"""
    
    def __init__(self, user):
        super().__init__()
        self.user = user
        self.db = DatabaseConnection()
        self.auth = Authentication()
        self.init_ui()
        self.load_products()
    
    def init_ui(self):
        """Initialize the UI components"""
        main_layout = QVBoxLayout(self)
        
        # Create tab widget for inventory sections
        tab_widget = QTabWidget()
        
        # Products tab
        products_tab = QWidget()
        products_layout = QVBoxLayout(products_tab)
        
        # Search and filter section
        search_frame = QFrame()
        search_frame.setFrameShape(QFrame.StyledPanel)
        search_layout = QHBoxLayout(search_frame)
        
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Search products...")
        self.search_input.textChanged.connect(self.filter_products)
        search_layout.addWidget(self.search_input)
        
        self.category_filter = QComboBox()
        self.category_filter.addItem("All Categories", None)
        self.load_category_filter()
        self.category_filter.currentIndexChanged.connect(self.filter_products)
        search_layout.addWidget(self.category_filter)
        
        # Medication Type filter
        self.type_filter = QComboBox()
        self.type_filter.addItem("All Types", None)
        self.type_filter.addItem("Branded", "branded")
        self.type_filter.addItem("Generic", "generic")
        self.type_filter.currentIndexChanged.connect(self.filter_products)
        search_layout.addWidget(self.type_filter)
        
        self.stock_filter = QComboBox()
        self.stock_filter.addItem("All Stock Levels", None)
        self.stock_filter.addItem("Low Stock", "low")
        self.stock_filter.addItem("In Stock", "in")
        self.stock_filter.addItem("Out of Stock", "out")
        self.stock_filter.currentIndexChanged.connect(self.filter_products)
        search_layout.addWidget(self.stock_filter)
        
        self.expiry_filter = QComboBox()
        self.expiry_filter.addItem("All Expiry Dates", None)
        self.expiry_filter.addItem("Expired", "expired")
        self.expiry_filter.addItem("Expiring Soon (30 days)", "soon")
        self.expiry_filter.addItem("Valid", "valid")
        self.expiry_filter.currentIndexChanged.connect(self.filter_products)
        search_layout.addWidget(self.expiry_filter)
        
        # Status filter
        self.status_filter = QComboBox()
        self.status_filter.addItem("All Products", None)
        self.status_filter.addItem("Active Only", "active")
        self.status_filter.addItem("Inactive Only", "inactive")
        self.status_filter.currentIndexChanged.connect(self.filter_products)
        search_layout.addWidget(self.status_filter)
        
        products_layout.addWidget(search_frame)
        
        # Products table - UPDATED with new columns
        self.products_table = QTableWidget()
        self.products_table.setColumnCount(11)
        self.products_table.setHorizontalHeaderLabels([
            "ID", "Name", "Description", "Type", "Category", "Unit", "Price", "Stock", "Expiry", "Status", "Actions"
        ])
        self.products_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.products_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.products_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.products_table.setEditTriggers(QTableWidget.NoEditTriggers)
        products_layout.addWidget(self.products_table)
        
        # Buttons for product management
        buttons_layout = QHBoxLayout()
        
        add_product_btn = QPushButton("Add Product")
        add_product_btn.setIcon(QIcon("resources/icons/add.png"))
        add_product_btn.clicked.connect(self.add_product)
        buttons_layout.addWidget(add_product_btn)
        
        edit_category_btn = QPushButton("Manage Categories")
        edit_category_btn.setIcon(QIcon("resources/icons/category.png"))
        edit_category_btn.clicked.connect(self.manage_categories)
        buttons_layout.addWidget(edit_category_btn)
        
        export_btn = QPushButton("Export Inventory")
        export_btn.setIcon(QIcon("resources/icons/export.png"))
        export_btn.clicked.connect(self.export_inventory)
        buttons_layout.addWidget(export_btn)
        
        refresh_btn = QPushButton("Refresh")
        refresh_btn.setIcon(QIcon("resources/icons/refresh.png"))
        refresh_btn.clicked.connect(self.load_products)
        buttons_layout.addWidget(refresh_btn)
        
        products_layout.addLayout(buttons_layout)
        
        # Add products tab
        tab_widget.addTab(products_tab, "Products")
        
        # Categories tab
        categories_tab = QWidget()
        categories_layout = QVBoxLayout(categories_tab)
        
        # Categories tree
        self.categories_tree = QTreeWidget()
        self.categories_tree.setHeaderLabels(["Category", "Products", "Actions"])
        self.categories_tree.header().setSectionResizeMode(0, QHeaderView.Stretch)
        categories_layout.addWidget(self.categories_tree)
        
        # Buttons for category management
        cat_buttons_layout = QHBoxLayout()
        
        add_category_btn = QPushButton("Add Category")
        add_category_btn.setIcon(QIcon("resources/icons/add.png"))
        add_category_btn.clicked.connect(self.add_category)
        cat_buttons_layout.addWidget(add_category_btn)
        
        refresh_cat_btn = QPushButton("Refresh")
        refresh_cat_btn.setIcon(QIcon("resources/icons/refresh.png"))
        refresh_cat_btn.clicked.connect(self.load_categories)
        cat_buttons_layout.addWidget(refresh_cat_btn)
        
        categories_layout.addLayout(cat_buttons_layout)
        
        # Add categories tab
        tab_widget.addTab(categories_tab, "Categories")
        
        # Stock alerts tab
        alerts_tab = QWidget()
        alerts_layout = QVBoxLayout(alerts_tab)
        
        # Low stock table
        self.low_stock_table = QTableWidget()
        self.low_stock_table.setColumnCount(6)
        self.low_stock_table.setHorizontalHeaderLabels([
            "Product Name", "Type", "Category", "Current Stock", "Reorder Level", "Actions"
        ])
        self.low_stock_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.low_stock_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        alerts_layout.addWidget(QLabel("Low Stock Items"))
        alerts_layout.addWidget(self.low_stock_table)
        
        # Expiring products table
        self.expiring_table = QTableWidget()
        self.expiring_table.setColumnCount(6)
        self.expiring_table.setHorizontalHeaderLabels([
            "Product Name", "Type", "Category", "Expiry Date", "Days Remaining", "Actions"
        ])
        self.expiring_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.expiring_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        alerts_layout.addWidget(QLabel("Expiring Soon or Expired Items"))
        alerts_layout.addWidget(self.expiring_table)
        
        # Add alerts tab
        tab_widget.addTab(alerts_tab, "Stock Alerts")
        
        # Add tab widget to main layout
        main_layout.addWidget(tab_widget)
        
        # Connect tab changed signal
        tab_widget.currentChanged.connect(self.tab_changed)
    
    def tab_changed(self, index):
        """Handle tab changes"""
        if index == 0:
            self.load_products()
        elif index == 1:
            self.load_categories()
        elif index == 2:
            self.load_alerts()
    
    def load_products(self):
        """Load products from database into table"""
        try:
            # Check if the database has the new medication fields
            connection = self.db.get_connection()
            cursor = connection.cursor()
            
            try:
                cursor.execute("""
                    SELECT column_name 
                    FROM information_schema.columns 
                    WHERE table_name = 'products' AND column_name = 'is_generic'
                """)
                has_med_fields = cursor.fetchone() is not None
            except:
                has_med_fields = False
            finally:
                cursor.close()
                self.db.release_connection(connection)
            
            # Update query based on whether the fields exist
            if has_med_fields:
                query = """
                    SELECT p.product_id, p.product_name, p.description, p.is_generic, c.name, p.unit_measurement,
                           p.unit_price, p.stock_quantity, p.expiry_date, p.reorder_level, p.is_active
                    FROM products p
                    LEFT JOIN categories c ON p.category_id = c.category_id
                    ORDER BY p.is_active DESC, p.product_name
                """
            else:
                query = """
                    SELECT p.product_id, p.product_name, p.description, NULL as is_generic, c.name, NULL as unit_measurement,
                           p.unit_price, p.stock_quantity, p.expiry_date, p.reorder_level, p.is_active
                    FROM products p
                    LEFT JOIN categories c ON p.category_id = c.category_id
                    ORDER BY p.is_active DESC, p.product_name
                """
            
            products = self.db.execute_query(query, fetchall=True)
            
            self.products_table.setRowCount(0)
            
            for row_idx, product in enumerate(products):
                self.products_table.insertRow(row_idx)
                
                product_id = product[0]
                name = product[1]
                description = product[2] or ""
                is_generic = product[3]
                category = product[4] or "Uncategorized"
                unit_measurement = product[5] or ""
                unit_price = product[6]
                stock_qty = product[7]
                expiry_date = product[8]
                reorder_level = product[9]
                is_active = product[10]
                
                # Product ID
                self.products_table.setItem(row_idx, 0, QTableWidgetItem(str(product_id)))
                
                # Product Name
                product_name_item = QTableWidgetItem(name)
                if not is_active:  # If not active
                    product_name_item.setForeground(QColor("#888888"))  # Grey text for inactive
                    font = product_name_item.font()
                    font.setStrikeOut(True)
                    product_name_item.setFont(font)
                self.products_table.setItem(row_idx, 1, product_name_item)
                
                # Description
                description_item = QTableWidgetItem(description)
                if not is_active:  # If not active
                    description_item.setForeground(QColor("#888888"))  # Grey text for inactive
                self.products_table.setItem(row_idx, 2, description_item)
                
                # Medication Type (Branded/Generic)
                type_text = "Generic" if is_generic else "Branded"
                self.products_table.setItem(row_idx, 3, QTableWidgetItem(type_text))
                
                # Category
                self.products_table.setItem(row_idx, 4, QTableWidgetItem(category))
                
                # Unit Measurement
                self.products_table.setItem(row_idx, 5, QTableWidgetItem(unit_measurement))
                
                # Unit Price
                price_item = QTableWidgetItem(f"₱{float(unit_price):.2f}")
                price_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.products_table.setItem(row_idx, 6, price_item)
                
                # Stock Quantity
                stock_item = QTableWidgetItem(str(stock_qty))
                stock_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                
                if stock_qty <= 0:
                    stock_item.setBackground(QColor("#FFCCCC"))  # Light red for out of stock
                elif stock_qty < reorder_level:
                    stock_item.setBackground(QColor("#FFFFCC"))  # Light yellow for low stock
                
                self.products_table.setItem(row_idx, 7, stock_item)
                
                # Expiry Date
                if expiry_date:
                    days_to_expiry = (expiry_date - QDate.currentDate().toPyDate()).days
                    expiry_item = QTableWidgetItem(expiry_date.strftime("%Y-%m-%d"))
                    
                    if days_to_expiry < 0:
                        expiry_item.setBackground(QColor("#FF9999"))  # Red for expired
                    elif days_to_expiry < 30:
                        expiry_item.setBackground(QColor("#FFCC99"))  # Orange for expiring soon
                    
                    self.products_table.setItem(row_idx, 8, expiry_item)
                else:
                    self.products_table.setItem(row_idx, 8, QTableWidgetItem("N/A"))
                
                # Status
                if not is_active:
                    status = "Inactive"
                    status_color = "#DDDDDD"  # Grey for inactive
                elif stock_qty <= 0:
                    status = "Out of Stock"
                    status_color = "#FFCCCC"  # Light red
                elif stock_qty < reorder_level:
                    status = "Low Stock"
                    status_color = "#FFFFCC"  # Light yellow
                elif expiry_date and (expiry_date - QDate.currentDate().toPyDate()).days < 0:
                    status = "Expired"
                    status_color = "#FF9999"  # Red
                elif expiry_date and (expiry_date - QDate.currentDate().toPyDate()).days < 30:
                    status = "Expiring Soon"
                    status_color = "#FFCC99"  # Orange
                else:
                    status = "OK"
                    status_color = "#CCFFCC"  # Light green
                
                status_item = QTableWidgetItem(status)
                status_item.setBackground(QColor(status_color))
                self.products_table.setItem(row_idx, 9, status_item)
                
                # Actions
                actions_widget = QWidget()
                actions_layout = QHBoxLayout(actions_widget)
                actions_layout.setContentsMargins(2, 2, 2, 2)
                actions_layout.setSpacing(2)
                
                edit_btn = QPushButton()
                edit_btn.setIcon(QIcon("resources/icons/edit.png"))
                edit_btn.setToolTip("Edit Product")
                edit_btn.setMaximumWidth(30)
                edit_btn.clicked.connect(lambda _, pid=product_id: self.edit_product(pid))
                actions_layout.addWidget(edit_btn)
                
                delete_btn = QPushButton()
                delete_btn.setIcon(QIcon("resources/icons/delete.png"))
                delete_btn.setToolTip("Delete Product")
                delete_btn.setMaximumWidth(30)
                delete_btn.clicked.connect(lambda _, pid=product_id, name=name: self.delete_product(pid, name))
                actions_layout.addWidget(delete_btn)
                
                self.products_table.setCellWidget(row_idx, 10, actions_widget)
            
            self.products_table.resizeColumnsToContents()
            self.products_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load products: {str(e)}")
            import traceback
            print(traceback.format_exc())
    
    def load_category_filter(self):
        """Load categories into filter dropdown"""
        try:
            query = "SELECT category_id, name FROM categories ORDER BY name"
            categories = self.db.execute_query(query, fetchall=True)
            
            for category_id, name in categories:
                self.category_filter.addItem(name, category_id)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load categories: {str(e)}")
    
    def filter_products(self):
        """Filter products based on search criteria"""
        search_text = self.search_input.text().lower()
        category_id = self.category_filter.currentData()
        type_filter = self.type_filter.currentData()
        stock_filter = self.stock_filter.currentData()
        expiry_filter = self.expiry_filter.currentData()
        status_filter = self.status_filter.currentData()
        
        for row in range(self.products_table.rowCount()):
            show_row = True
            
            # Filter by product name
            product_name = self.products_table.item(row, 1).text().lower()
            if search_text and search_text not in product_name:
                show_row = False
            
            # Filter by medication type (branded/generic)
            if type_filter:
                type_text = self.products_table.item(row, 3).text().lower()
                if type_filter == "branded" and type_text != "branded":
                    show_row = False
                elif type_filter == "generic" and type_text != "generic":
                    show_row = False
            
            # Filter by category
            if category_id is not None:
                category_name = self.products_table.item(row, 4).text()
                category_match = False
                
                for i in range(self.category_filter.count()):
                    if (self.category_filter.itemData(i) == category_id and 
                        self.category_filter.itemText(i) == category_name):
                        category_match = True
                        break
                
                if not category_match:
                    show_row = False
            
            # Filter by stock level
            if stock_filter:
                stock_qty = int(self.products_table.item(row, 7).text())
                reorder_level = 10  # Default reorder level
                
                if stock_filter == "low" and stock_qty > reorder_level:
                    show_row = False
                elif stock_filter == "out" and stock_qty > 0:
                    show_row = False
                elif stock_filter == "in" and stock_qty <= 0:
                    show_row = False
            
            # Filter by expiry date
            if expiry_filter and self.products_table.item(row, 8).text() != "N/A":
                expiry_date = QDate.fromString(self.products_table.item(row, 8).text(), "yyyy-MM-dd")
                days_to_expiry = QDate.currentDate().daysTo(expiry_date)
                
                if expiry_filter == "expired" and days_to_expiry >= 0:
                    show_row = False
                elif expiry_filter == "soon" and (days_to_expiry < 0 or days_to_expiry > 30):
                    show_row = False
                elif expiry_filter == "valid" and days_to_expiry < 0:
                    show_row = False
            
            # Filter by status - use column 9 which has the status text
            if status_filter is not None:
                status_text = self.products_table.item(row, 9).text()
                is_active = status_text != "Inactive"  # Any status except "Inactive" is considered active
                
                if status_filter == "active" and not is_active:
                    show_row = False
                elif status_filter == "inactive" and is_active:
                    show_row = False
        
            self.products_table.setRowHidden(row, not show_row)
    
    def add_product(self):
        """Add a new product"""
        dialog = ProductDialog(self)
        if dialog.exec_():
            self.load_products()
            self.load_alerts()
    
    def edit_product(self, product_id):
        """Edit an existing product"""
        dialog = ProductDialog(self, product_id)
        if dialog.exec_():
            self.load_products()
            self.load_alerts()
    
    def delete_product(self, product_id, product_name):
        """Delete a product"""
        reply = QMessageBox.question(
            self, 
            "Confirm Delete", 
            f"Are you sure you want to delete '{product_name}'?\nThis action cannot be undone.",
            QMessageBox.Yes | QMessageBox.No, 
            QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            try:
                # Check if product is used in any sales
                check_query = """
                    SELECT COUNT(*) FROM sale_items 
                    WHERE product_id = %s
                """
                count = self.db.execute_query(check_query, (product_id,), fetchone=True)[0]
                
                if count > 0:
                    deactivate_reply = QMessageBox.question(
                        self, 
                        "Cannot Delete - Deactivate Instead?", 
                        f"Cannot delete '{product_name}' because it is used in {count} sales.\n\nWould you like to deactivate it instead? Deactivated products will not appear in sales but will remain in reports.",
                        QMessageBox.Yes | QMessageBox.No, 
                        QMessageBox.Yes
                    )
                    
                    if deactivate_reply == QMessageBox.Yes:
                        # Deactivate the product
                        query = "UPDATE products SET is_active = FALSE WHERE product_id = %s"
                        self.db.execute_query(query, (product_id,))
                        
                        # Log activity
                        self.auth.log_activity(
                            self.user['user_id'],
                            "update",
                            "products",
                            product_id,
                            f"Deactivated product: {product_name}"
                        )
                        
                        QMessageBox.information(self, "Success", f"Product '{product_name}' deactivated successfully.")
                        self.load_products()
                        self.load_alerts()
                    
                    return
                
                # Delete the product
                query = "DELETE FROM products WHERE product_id = %s"
                self.db.execute_query(query, (product_id,))
                
                # Log activity
                self.auth.log_activity(
                    self.user['user_id'],
                    "delete",
                    "products",
                    product_id,
                    f"Deleted product: {product_name}"
                )
                
                QMessageBox.information(self, "Success", f"Product '{product_name}' deleted successfully.")
                self.load_products()
                self.load_alerts()
                
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to delete product: {str(e)}")
                import traceback
                print(traceback.format_exc())
    
    def load_categories(self):
        """Load categories into the tree widget"""
        try:
            self.categories_tree.clear()
            
            # Get all categories with product count
            query = """
                SELECT c.category_id, c.name, c.description, COUNT(p.product_id) AS product_count
                FROM categories c
                LEFT JOIN products p ON c.category_id = p.category_id
                GROUP BY c.category_id, c.name, c.description
                ORDER BY c.name
            """
            categories = self.db.execute_query(query, fetchall=True)
            
            for category in categories:
                category_id, name, description, product_count = category
                
                # Create category item
                item = QTreeWidgetItem(self.categories_tree)
                item.setText(0, name)
                item.setText(1, str(product_count))
                item.setData(0, Qt.UserRole, category_id)
                
                # Add tooltip with description
                if description:
                    item.setToolTip(0, description)
                
                # Add action buttons
                actions_widget = QWidget()
                actions_layout = QHBoxLayout(actions_widget)
                actions_layout.setContentsMargins(2, 2, 2, 2)
                actions_layout.setSpacing(2)
                
                edit_btn = QPushButton()
                edit_btn.setIcon(QIcon("resources/icons/edit.png"))
                edit_btn.setToolTip("Edit Category")
                edit_btn.setMaximumWidth(30)
                edit_btn.clicked.connect(lambda _, cid=category_id: self.edit_category(cid))
                actions_layout.addWidget(edit_btn)
                
                delete_btn = QPushButton()
                delete_btn.setIcon(QIcon("resources/icons/delete.png"))
                delete_btn.setToolTip("Delete Category")
                delete_btn.setMaximumWidth(30)
                delete_btn.clicked.connect(lambda _, cid=category_id, cname=name: self.delete_category(cid, cname))
                actions_layout.addWidget(delete_btn)
                
                self.categories_tree.setItemWidget(item, 2, actions_widget)
            
            self.categories_tree.expandAll()
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load categories: {str(e)}")
            import traceback
            print(traceback.format_exc())
    
    def add_category(self):
        """Add a new category"""
        dialog = CategoryDialog(self)
        if dialog.exec_():
            self.load_categories()
            self.load_category_filter()
    
    def edit_category(self, category_id):
        """Edit an existing category"""
        dialog = CategoryDialog(self, category_id)
        if dialog.exec_():
            self.load_categories()
            self.load_category_filter()
    
    def delete_category(self, category_id, category_name):
        """Delete a category"""
        reply = QMessageBox.question(
            self, 
            "Confirm Delete", 
            f"Are you sure you want to delete category '{category_name}'?\nThis action cannot be undone.",
            QMessageBox.Yes | QMessageBox.No, 
            QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            try:
                # Check if category has products
                check_query = """
                    SELECT COUNT(*) FROM products 
                    WHERE category_id = %s
                """
                count = self.db.execute_query(check_query, (category_id,), fetchone=True)[0]
                
                if count > 0:
                    QMessageBox.warning(
                        self, 
                        "Cannot Delete", 
                        f"Cannot delete '{category_name}' because it contains {count} products.\nRemove or reassign these products first."
                    )
                    return
                
                # Delete the category
                query = "DELETE FROM categories WHERE category_id = %s"
                self.db.execute_query(query, (category_id,))
                
                # Log activity
                self.auth.log_activity(
                    self.user['user_id'],
                    "delete",
                    "categories",
                    category_id,
                    f"Deleted category: {category_name}"
                )
                
                QMessageBox.information(self, "Success", f"Category '{category_name}' deleted successfully.")
                self.load_categories()
                self.load_category_filter()
                
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to delete category: {str(e)}")
                import traceback
                print(traceback.format_exc())
    
    def manage_categories(self):
        """Open the categories tab"""
        # Find the QTabWidget directly within this widget's children
        for widget in self.children():
            if isinstance(widget, QTabWidget):
                widget.setCurrentIndex(1)  # Switch to Categories tab
                return
        
        # If we get here, we couldn't find the tab widget
        QMessageBox.information(self, "Categories", 
            "Please click on the 'Categories' tab to manage categories.")
    
    def load_alerts(self):
        """Load stock alerts (low stock and expiring products)"""
        self.load_low_stock()
        self.load_expiring_products()
    
    def load_low_stock(self):
        """Load low stock items"""
        try:
            # Check if the database has the new medication fields
            connection = self.db.get_connection()
            cursor = connection.cursor()
            
            try:
                cursor.execute("""
                    SELECT column_name 
                    FROM information_schema.columns 
                    WHERE table_name = 'products' AND column_name = 'is_generic'
                """)
                has_med_fields = cursor.fetchone() is not None
            except:
                has_med_fields = False
            finally:
                cursor.close()
                self.db.release_connection(connection)
            
            # Update query based on whether the fields exist
            if has_med_fields:
                query = """
                    SELECT p.product_id, p.product_name, p.description, p.is_generic, c.name, 
                           p.stock_quantity, p.reorder_level
                    FROM products p
                    LEFT JOIN categories c ON p.category_id = c.category_id
                    WHERE p.stock_quantity <= p.reorder_level AND p.is_active = TRUE
                    ORDER BY (p.reorder_level - p.stock_quantity) DESC, p.product_name
                """
            else:
                query = """
                    SELECT p.product_id, p.product_name, p.description, NULL as is_generic, c.name, 
                           p.stock_quantity, p.reorder_level
                    FROM products p
                    LEFT JOIN categories c ON p.category_id = c.category_id
                    WHERE p.stock_quantity <= p.reorder_level AND p.is_active = TRUE
                    ORDER BY (p.reorder_level - p.stock_quantity) DESC, p.product_name
                """
            
            low_stock = self.db.execute_query(query, fetchall=True)
            
            self.low_stock_table.setRowCount(0)
            
            for row_idx, product in enumerate(low_stock):
                self.low_stock_table.insertRow(row_idx)
                
                product_id, product_name, description, is_generic, category_name, stock_qty, reorder_level = product
                
                # Product Name
                self.low_stock_table.setItem(row_idx, 0, QTableWidgetItem(product_name))
                
                # Medication Type
                type_text = "Generic" if is_generic else "Branded"
                self.low_stock_table.setItem(row_idx, 1, QTableWidgetItem(type_text))
                
                # Category
                self.low_stock_table.setItem(row_idx, 2, QTableWidgetItem(category_name or "Uncategorized"))
                
                # Current Stock
                stock_item = QTableWidgetItem(str(stock_qty))
                stock_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                
                if stock_qty <= 0:
                    stock_item.setBackground(QColor("#FFCCCC"))  # Light red for out of stock
                else:
                    stock_item.setBackground(QColor("#FFFFCC"))  # Light yellow for low stock
                
                self.low_stock_table.setItem(row_idx, 3, stock_item)
                
                # Reorder Level
                reorder_item = QTableWidgetItem(str(reorder_level))
                reorder_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.low_stock_table.setItem(row_idx, 4, reorder_item)
                
                # Actions
                actions_widget = QWidget()
                actions_layout = QHBoxLayout(actions_widget)
                actions_layout.setContentsMargins(2, 2, 2, 2)
                actions_layout.setSpacing(2)
                
                edit_btn = QPushButton()
                edit_btn.setIcon(QIcon("resources/icons/edit.png"))
                edit_btn.setToolTip("Edit Product")
                edit_btn.setMaximumWidth(30)
                edit_btn.clicked.connect(lambda _, pid=product_id: self.edit_product(pid))
                actions_layout.addWidget(edit_btn)
                
                self.low_stock_table.setCellWidget(row_idx, 5, actions_widget)
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load low stock items: {str(e)}")
            import traceback
            print(traceback.format_exc())
    
    def load_expiring_products(self):
        """Load expiring products"""
        try:
            # Check if the database has the new medication fields
            connection = self.db.get_connection()
            cursor = connection.cursor()
            
            try:
                cursor.execute("""
                    SELECT column_name 
                    FROM information_schema.columns 
                    WHERE table_name = 'products' AND column_name = 'is_generic'
                """)
                has_med_fields = cursor.fetchone() is not None
            except:
                has_med_fields = False
            finally:
                cursor.close()
                self.db.release_connection(connection)
            
            # Update query based on whether the fields exist
            if has_med_fields:
                query = """
                    SELECT p.product_id, p.product_name, p.description, p.is_generic, c.name, p.expiry_date
                    FROM products p
                    LEFT JOIN categories c ON p.category_id = c.category_id
                    WHERE p.expiry_date IS NOT NULL AND p.expiry_date <= %s AND p.is_active = TRUE
                    ORDER BY p.expiry_date, p.product_name
                """
            else:
                query = """
                    SELECT p.product_id, p.product_name, p.description, NULL as is_generic, c.name, p.expiry_date
                    FROM products p
                    LEFT JOIN categories c ON p.category_id = c.category_id
                    WHERE p.expiry_date IS NOT NULL AND p.expiry_date <= %s AND p.is_active = TRUE
                    ORDER BY p.expiry_date, p.product_name
                """
            
            thirty_days_later = QDate.currentDate().addDays(30).toPyDate()
            expiring = self.db.execute_query(query, (thirty_days_later,), fetchall=True)
            
            self.expiring_table.setRowCount(0)
            
            for row_idx, product in enumerate(expiring):
                self.expiring_table.insertRow(row_idx)
                
                product_id, product_name, description, is_generic, category_name, expiry_date = product
                
                # Product Name
                self.expiring_table.setItem(row_idx, 0, QTableWidgetItem(product_name))
                
                # Medication Type
                type_text = "Generic" if is_generic else "Branded"
                self.expiring_table.setItem(row_idx, 1, QTableWidgetItem(type_text))
                
                # Category
                self.expiring_table.setItem(row_idx, 2, QTableWidgetItem(category_name or "Uncategorized"))
                
                # Expiry Date
                expiry_item = QTableWidgetItem(expiry_date.strftime("%Y-%m-%d"))
                days_to_expiry = (expiry_date - QDate.currentDate().toPyDate()).days
                
                if days_to_expiry < 0:
                    expiry_item.setBackground(QColor("#FF9999"))  # Red for expired
                else:
                    expiry_item.setBackground(QColor("#FFCC99"))  # Orange for expiring soon
                
                self.expiring_table.setItem(row_idx, 3, expiry_item)
                
                # Days Remaining
                days_item = QTableWidgetItem(str(days_to_expiry) if days_to_expiry >= 0 else "Expired")
                days_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                
                if days_to_expiry < 0:
                    days_item.setBackground(QColor("#FF9999"))  # Red for expired
                else:
                    days_item.setBackground(QColor("#FFCC99"))  # Orange for expiring soon
                
                self.expiring_table.setItem(row_idx, 4, days_item)
                
                # Actions
                actions_widget = QWidget()
                actions_layout = QHBoxLayout(actions_widget)
                actions_layout.setContentsMargins(2, 2, 2, 2)
                actions_layout.setSpacing(2)
                
                edit_btn = QPushButton()
                edit_btn.setIcon(QIcon("resources/icons/edit.png"))
                edit_btn.setToolTip("Edit Product")
                edit_btn.setMaximumWidth(30)
                edit_btn.clicked.connect(lambda _, pid=product_id: self.edit_product(pid))
                actions_layout.addWidget(edit_btn)
                
                self.expiring_table.setCellWidget(row_idx, 5, actions_widget)
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load expiring products: {str(e)}")
            import traceback
            print(traceback.format_exc())
    
    def export_inventory(self):
        """Export inventory to Excel"""
        try:
            import pandas as pd
            from datetime import datetime
            from PyQt5.QtWidgets import QFileDialog
            
            # Get save file location
            filename, _ = QFileDialog.getSaveFileName(
                self,
                "Export Inventory",
                f"inventory_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                "Excel Files (*.xlsx)"
            )
            
            if not filename:
                return
            
            # Check if the database has the new medication fields
            connection = self.db.get_connection()
            cursor = connection.cursor()
            
            try:
                cursor.execute("""
                    SELECT column_name 
                    FROM information_schema.columns 
                    WHERE table_name = 'products' AND column_name = 'is_generic'
                """)
                has_med_fields = cursor.fetchone() is not None
            except:
                has_med_fields = False
            finally:
                cursor.close()
                self.db.release_connection(connection)
            
            # Fetch data from database
            if has_med_fields:
                query = """
                    SELECT p.product_id, p.product_name, p.is_generic, c.name as category, 
                           p.unit_measurement, p.description, p.unit_price, p.cost_price, 
                           p.stock_quantity, p.expiry_date, p.reorder_level, s.name as supplier, 
                           p.is_active
                    FROM products p
                    LEFT JOIN categories c ON p.category_id = c.category_id
                    LEFT JOIN suppliers s ON p.supplier_id = s.supplier_id
                    ORDER BY p.product_name
                """
                
                products = self.db.execute_query(query, fetchall=True)
                
                # Convert to pandas dataframe
                columns = [
                    "Product ID", "Product Name", "Is Generic", "Category", "Unit Measurement",
                    "Description", "Unit Price", "Cost Price", "Stock Quantity", 
                    "Expiry Date", "Reorder Level", "Supplier", "Is Active"
                ]
                
                df = pd.DataFrame(products, columns=columns)
                
                # Convert boolean columns to text
                df['Type'] = df['Is Generic'].apply(lambda x: 'Generic' if x else 'Branded')
                df['Status'] = df['Is Active'].apply(lambda x: 'Active' if x else 'Inactive')
                
                # Drop original boolean columns
                df = df.drop(['Is Generic', 'Is Active'], axis=1)
                                # Reorder columns for better readability
                df = df[['Product ID', 'Product Name', 'Type', 'Category', 'Unit Measurement',
                        'Description', 'Unit Price', 'Cost Price', 'Stock Quantity',
                        'Expiry Date', 'Reorder Level', 'Supplier', 'Status']]
            else:
                # Original query without medication details
                query = """
                    SELECT p.product_id, p.product_name, c.name as category, p.description,
                           p.unit_price, p.cost_price, p.stock_quantity, p.expiry_date,
                           p.reorder_level, s.name as supplier, p.is_active
                    FROM products p
                    LEFT JOIN categories c ON p.category_id = c.category_id
                    LEFT JOIN suppliers s ON p.supplier_id = s.supplier_id
                    ORDER BY p.product_name
                """
                products = self.db.execute_query(query, fetchall=True)
                
                # Convert to pandas dataframe
                columns = [
                    "Product ID", "Product Name", "Category", "Description",
                    "Unit Price", "Cost Price", "Stock Quantity", "Expiry Date",
                    "Reorder Level", "Supplier", "Is Active"
                ]
                
                df = pd.DataFrame(products, columns=columns)
                
                # Convert boolean column to text
                df['Status'] = df['Is Active'].apply(lambda x: 'Active' if x else 'Inactive')
                
                # Drop original boolean column
                df = df.drop('Is Active', axis=1)
                
                # Add Type and Unit Measurement columns with empty values
                df.insert(2, 'Type', 'Branded')  # Default to Branded
                df.insert(4, 'Unit Measurement', '')
                
                # Reorder columns for better readability
                df = df[['Product ID', 'Product Name', 'Type', 'Category', 'Unit Measurement',
                        'Description', 'Unit Price', 'Cost Price', 'Stock Quantity',
                        'Expiry Date', 'Reorder Level', 'Supplier', 'Status']]
            
            # Format date columns
            if 'Expiry Date' in df.columns:
                df['Expiry Date'] = df['Expiry Date'].apply(
                    lambda x: x.strftime('%Y-%m-%d') if pd.notnull(x) else 'N/A'
                )
            
            # Format numeric columns
            for col in ['Unit Price', 'Cost Price']:
                if col in df.columns:
                    df[col] = df[col].apply(lambda x: f"₱{float(x):.2f}" if pd.notnull(x) else '₱0.00')
            
            # Add export metadata
            metadata_df = pd.DataFrame([
                {'Metadata': 'Export Date', 'Value': datetime.now().strftime('%Y-%m-%d %H:%M:%S')},
                {'Metadata': 'Exported By', 'Value': self.user['username'] if 'username' in self.user else 'System'},
                {'Metadata': 'Total Products', 'Value': len(df)},
                {'Metadata': 'Active Products', 'Value': len(df[df['Status'] == 'Active'])},
                {'Metadata': 'Low Stock Items', 'Value': len(df[(df['Status'] == 'Active') & 
                                                              (df['Stock Quantity'].astype(float) <= df['Reorder Level'].astype(float))])}
            ])
            
            # Write both dataframes to separate sheets in the Excel file
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # Add metadata sheet
                metadata_df.to_excel(writer, sheet_name='Export Info', index=False)
                
                # Add products sheet
                df.to_excel(writer, sheet_name='Products', index=False)
                
                # Auto-adjust column widths in both sheets
                for sheet_name in writer.sheets:
                    worksheet = writer.sheets[sheet_name]
                    for idx, col in enumerate(worksheet.columns, 1):
                        # Calculate max column width based on content
                        max_width = 0
                        for cell in col:
                            if cell.value:
                                # Add padding to the length
                                cell_width = len(str(cell.value)) + 2
                                max_width = max(max_width, cell_width)
                        # Set column width with a maximum reasonable value
                        worksheet.column_dimensions[worksheet.cell(row=1, column=idx).column_letter].width = min(max_width, 40)
            
            # Format the Excel file with header styling using openpyxl
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            wb = load_workbook(filename)
            
            # Style the Products sheet
            ws = wb['Products']
            
            header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='E74C3C', end_color='E74C3C', fill_type='solid')
            
            # Style the headers
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Style the metadata sheet
            ws_meta = wb['Export Info']
            
            for cell in ws_meta[1]:
                cell.font = header_font
                cell.fill = header_fill
            
            metadata_border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
            
            # Add borders to all metadata cells
            for row in ws_meta.iter_rows(min_row=1, max_row=len(metadata_df) + 1):
                for cell in row:
                    cell.border = metadata_border
            
            # Save the styled workbook
            wb.save(filename)
            
            # Log activity
            self.auth.log_activity(
                self.user['user_id'],
                "export",
                "products",
                None,
                f"Exported inventory to Excel: {filename}"
            )
            
            QMessageBox.information(
                self,
                "Export Successful",
                f"Inventory exported successfully to:\n{filename}"
            )
            
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed to export inventory: {str(e)}")
            import traceback
            print(traceback.format_exc())


# SQL script to add the medication type and unit columns to the database
"""
-- Add medication detail columns to products table if they don't exist
ALTER TABLE products ADD COLUMN IF NOT EXISTS is_generic BOOLEAN DEFAULT FALSE;
ALTER TABLE products ADD COLUMN IF NOT EXISTS unit_measurement VARCHAR(50);

-- Add medication detail columns to sale_items table if they don't exist
ALTER TABLE sale_items ADD COLUMN IF NOT EXISTS is_generic BOOLEAN DEFAULT FALSE;
ALTER TABLE sale_items ADD COLUMN IF NOT EXISTS unit_measurement VARCHAR(50);

-- Update any existing NULL values to default values
UPDATE products SET is_generic = FALSE WHERE is_generic IS NULL;
UPDATE products SET unit_measurement = '' WHERE unit_measurement IS NULL;

-- Update existing records in sale_items to match their products
UPDATE sale_items si 
SET 
    is_generic = p.is_generic, 
    unit_measurement = p.unit_measurement
FROM products p
WHERE si.product_id = p.product_id;
"""