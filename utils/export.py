import pandas as pd
from PyQt5.QtWidgets import (QFileDialog, QMessageBox, QProgressDialog, 
                            QApplication, QDialog, QVBoxLayout, QHBoxLayout,
                            QLabel, QComboBox, QCheckBox, QPushButton, QGroupBox,
                            QRadioButton, QButtonGroup, QSpinBox, QLineEdit, QFrame)
from PyQt5.QtCore import QDate, Qt, QSize, QThread, pyqtSignal
from PyQt5.QtGui import QIcon, QFont, QColor, QPixmap
from datetime import datetime
import os
import time
import threading
import re

# PDF Generation
from reportlab.lib.pagesizes import letter, A4, landscape
from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle, Paragraph, 
                               Spacer, Image, PageBreak, NextPageTemplate)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch, cm
from reportlab.graphics.shapes import Drawing, Line
from reportlab.pdfgen import canvas

# Excel handling
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage


class ExportThread(QThread):
    """Thread for handling export operations without freezing the UI"""
    progress_update = pyqtSignal(int)
    finished = pyqtSignal(bool, str)
    
    def __init__(self, export_function, *args, **kwargs):
        super().__init__()
        self.export_function = export_function
        self.args = args
        self.kwargs = kwargs
    
    def run(self):
        try:
            # Simulate progress for operations that don't report progress
            for i in range(0, 101, 10):
                self.progress_update.emit(i)
                time.sleep(0.1)  # Small delay to show progress
            
            # Call the actual export function
            result, message = self.export_function(*self.args, **self.kwargs)
            self.finished.emit(result, message)
        except Exception as e:
            self.finished.emit(False, str(e))


class ExportOptionsDialog(QDialog):
    """Dialog for configuring export options"""
    
    def __init__(self, parent=None, export_type="excel", has_images=False):
        super().__init__(parent)
        self.export_type = export_type
        self.has_images = has_images
        self.setup_ui()
    
    def setup_ui(self):
        """Set up the dialog UI"""
        if self.export_type == "excel":
            self.setWindowTitle("Excel Export Options")
            self.setWindowIcon(QIcon("resources/icons/excel.png"))
        else:
            self.setWindowTitle("PDF Export Options")
            self.setWindowIcon(QIcon("resources/icons/pdf.png"))
        
        self.setMinimumWidth(400)
        
        main_layout = QVBoxLayout(self)
        
        # File options group
        file_group = QGroupBox("File Options")
        file_layout = QVBoxLayout(file_group)
        
        # Sheet name / Document title
        title_layout = QHBoxLayout()
        if self.export_type == "excel":
            title_label = QLabel("Sheet Name:")
            self.title_input = QLineEdit("Data")
        else:
            title_label = QLabel("Document Title:")
            self.title_input = QLineEdit("Pharmacy Management System Report")
        
        title_layout.addWidget(title_label)
        title_layout.addWidget(self.title_input)
        file_layout.addLayout(title_layout)
        
        # Include company info
        self.include_company_info = QCheckBox("Include Company Information")
        self.include_company_info.setChecked(True)
        file_layout.addWidget(self.include_company_info)
        
        # Include date/time
        self.include_datetime = QCheckBox("Include Date and Time")
        self.include_datetime.setChecked(True)
        file_layout.addWidget(self.include_datetime)
        
        main_layout.addWidget(file_group)
        
        # Styling options group
        style_group = QGroupBox("Styling Options")
        style_layout = QVBoxLayout(style_group)
        
        # Color scheme
        color_layout = QHBoxLayout()
        color_layout.addWidget(QLabel("Color Scheme:"))
        
        self.color_scheme = QComboBox()
        self.color_scheme.addItems(["Blue", "Green", "Red", "Purple", "Orange", "Grayscale"])
        color_layout.addWidget(self.color_scheme)
        style_layout.addLayout(color_layout)
        
        # Font size
        font_layout = QHBoxLayout()
        font_layout.addWidget(QLabel("Font Size:"))
        
        self.font_size = QSpinBox()
        self.font_size.setRange(8, 16)
        self.font_size.setValue(10)
        font_layout.addWidget(self.font_size)
        style_layout.addLayout(font_layout)
        
        # Paper size (PDF only)
        if self.export_type == "pdf":
            paper_layout = QHBoxLayout()
            paper_layout.addWidget(QLabel("Paper Size:"))
            
            self.paper_size = QComboBox()
            self.paper_size.addItems(["Letter", "A4", "Legal"])
            paper_layout.addWidget(self.paper_size)
            style_layout.addLayout(paper_layout)
            
            # Orientation
            orientation_layout = QHBoxLayout()
            orientation_layout.addWidget(QLabel("Orientation:"))
            
            self.orientation_group = QButtonGroup(self)
            
            self.portrait_radio = QRadioButton("Portrait")
            self.portrait_radio.setChecked(True)
            self.orientation_group.addButton(self.portrait_radio)
            orientation_layout.addWidget(self.portrait_radio)
            
            self.landscape_radio = QRadioButton("Landscape")
            self.orientation_group.addButton(self.landscape_radio)
            orientation_layout.addWidget(self.landscape_radio)
            
            style_layout.addLayout(orientation_layout)
        
        # Row striping
        self.alternate_rows = QCheckBox("Alternate Row Colors")
        self.alternate_rows.setChecked(True)
        style_layout.addWidget(self.alternate_rows)
        
        main_layout.addWidget(style_group)
        
        # Content options group
        content_group = QGroupBox("Content Options")
        content_layout = QVBoxLayout(content_group)
        
        # Include images (if applicable)
        if self.has_images:
            self.include_images = QCheckBox("Include Images")
            self.include_images.setChecked(True)
            content_layout.addWidget(self.include_images)
        
        # Include totals
        self.include_totals = QCheckBox("Include Row/Column Totals")
        self.include_totals.setChecked(False)
        content_layout.addWidget(self.include_totals)
        
        # Currency symbol
        currency_layout = QHBoxLayout()
        currency_layout.addWidget(QLabel("Currency Symbol:"))
        
        self.currency_symbol = QComboBox()
        self.currency_symbol.addItems(["₱ (Philippine Peso)", "$ (US Dollar)", "€ (Euro)", "£ (British Pound)"])
        currency_layout.addWidget(self.currency_symbol)
        content_layout.addLayout(currency_layout)
        
        # Auto-fit columns (Excel only)
        if self.export_type == "excel":
            self.autofit_columns = QCheckBox("Auto-fit Column Widths")
            self.autofit_columns.setChecked(True)
            content_layout.addWidget(self.autofit_columns)
            
            # Freeze header row
            self.freeze_header = QCheckBox("Freeze Header Row")
            self.freeze_header.setChecked(True)
            content_layout.addWidget(self.freeze_header)
        
        main_layout.addWidget(content_group)
        
        # Buttons
        button_layout = QHBoxLayout()
        
        reset_btn = QPushButton("Reset to Defaults")
        reset_btn.clicked.connect(self.reset_defaults)
        button_layout.addWidget(reset_btn)
        
        button_layout.addStretch()
        
        cancel_btn = QPushButton("Cancel")
        cancel_btn.clicked.connect(self.reject)
        button_layout.addWidget(cancel_btn)
        
        ok_btn = QPushButton("OK")
        ok_btn.clicked.connect(self.accept)
        ok_btn.setDefault(True)
        button_layout.addWidget(ok_btn)
        
        main_layout.addLayout(button_layout)
    
    def reset_defaults(self):
        """Reset all options to default values"""
        self.title_input.setText("Data" if self.export_type == "excel" else "Pharmacy Management System Report")
        self.include_company_info.setChecked(True)
        self.include_datetime.setChecked(True)
        self.color_scheme.setCurrentIndex(0)  # Blue
        self.font_size.setValue(10)
        self.currency_symbol.setCurrentIndex(0)  # Philippine Peso
        
        if self.export_type == "pdf":
            self.paper_size.setCurrentIndex(0)  # Letter
            self.portrait_radio.setChecked(True)
        
        self.alternate_rows.setChecked(True)
        
        if self.has_images:
            self.include_images.setChecked(True)
        
        self.include_totals.setChecked(False)
        
        if self.export_type == "excel":
            self.autofit_columns.setChecked(True)
            self.freeze_header.setChecked(True)
    
    def get_options(self):
        """Get all selected options as a dictionary"""
        options = {
            "title": self.title_input.text(),
            "include_company_info": self.include_company_info.isChecked(),
            "include_datetime": self.include_datetime.isChecked(),
            "color_scheme": self.color_scheme.currentText().lower(),
            "font_size": self.font_size.value(),
            "alternate_rows": self.alternate_rows.isChecked(),
            "include_totals": self.include_totals.isChecked(),
            "currency_symbol": self.currency_symbol.currentText().split()[0]
        }
        
        if self.has_images:
            options["include_images"] = self.include_images.isChecked()
        
        if self.export_type == "pdf":
            options["paper_size"] = self.paper_size.currentText()
            options["orientation"] = "portrait" if self.portrait_radio.isChecked() else "landscape"
        
        if self.export_type == "excel":
            options["autofit_columns"] = self.autofit_columns.isChecked()
            options["freeze_header"] = self.freeze_header.isChecked()
        
        return options


class ExportUtility:
    """Enhanced utility class for exporting data to various formats"""
    
    # Color schemes
    COLOR_SCHEMES = {
        "blue": {
            "header_fill": "4472C4",
            "header_font": "FFFFFF",
            "alt_row_fill": "D9E1F2",
            "border": "8EA9DB",
            "title_font": "2F5597"
        },
        "green": {
            "header_fill": "70AD47",
            "header_font": "FFFFFF",
            "alt_row_fill": "E2EFDA",
            "border": "A9D08E",
            "title_font": "538135"
        },
        "red": {
            "header_fill": "C00000",
            "header_font": "FFFFFF",
            "alt_row_fill": "F8CBAD",
            "border": "F4B084",
            "title_font": "943634"
        },
        "purple": {
            "header_fill": "7030A0",
            "header_font": "FFFFFF",
            "alt_row_fill": "E4DFEC",
            "border": "B4A7D6",
            "title_font": "632C8D"
        },
        "orange": {
            "header_fill": "ED7D31",
            "header_font": "FFFFFF",
            "alt_row_fill": "FCE4D6",
            "border": "F8CBAD",
            "title_font": "C55A11"
        },
        "grayscale": {
            "header_fill": "404040",
            "header_font": "FFFFFF",
            "alt_row_fill": "E7E6E6",
            "border": "BFBFBF",
            "title_font": "202020"
        }
    }
    
    @staticmethod
    def get_company_info():
        """Get company information from config file"""
        try:
            from configparser import ConfigParser
            config = ConfigParser()
            config.read('config.ini')
            
            company_info = {
                "name": config.get('application', 'company', fallback='Scorpions Pharmacy'),
                "logo_path": config.get('application', 'logo_path', fallback='resources/icons/pharmacy.png')
            }
            
            return company_info
        except Exception:
            # Default values if there's an error
            return {
                "name": "Scorpions Pharmacy",
                "logo_path": "resources/icons/pharmacy.png"
            }
    
    @staticmethod
    def replace_currency_symbol(value, currency_symbol="₱"):
        """Replace $ with the specified currency symbol in a value"""
        if isinstance(value, str):
            # Replace dollar signs with the specified currency symbol
            return value.replace("$", currency_symbol)
        return value
    
    @staticmethod
    def process_data_with_currency(data, headers, currency_symbol="₱"):
        """Process data to replace currency symbols"""
        # Create deep copies to avoid modifying original data
        processed_headers = list(headers)
        processed_data = []
        
        # Process headers (replace $ in column names)
        for i, header in enumerate(processed_headers):
            processed_headers[i] = ExportUtility.replace_currency_symbol(header, currency_symbol)
        
        # Process data rows
        for row in data:
            processed_row = []
            for cell in row:
                processed_row.append(ExportUtility.replace_currency_symbol(cell, currency_symbol))
            processed_data.append(processed_row)
        
        return processed_headers, processed_data
    
    @staticmethod
    def export_to_excel(parent, data, headers, filename_prefix, sheet_name=None, options=None):
        """
        Enhanced export to Excel file with styling options
        
        Parameters:
        - parent: Parent widget for file dialog
        - data: List of data rows
        - headers: List of column headers
        - filename_prefix: Prefix for the default filename
        - sheet_name: Name of the Excel sheet (optional)
        - options: Dictionary of export options (optional)
        
        Returns:
        - tuple: (bool: success, str: message)
        """
        try:
            # Show export options dialog if not provided
            if options is None:
                dialog = ExportOptionsDialog(parent, export_type="excel")
                if dialog.exec_() != QDialog.Accepted:
                    return False, "Export cancelled"
                
                options = dialog.get_options()
                sheet_name = options["title"]
            elif not sheet_name:
                sheet_name = "Data"
            
            # Process data to replace currency symbols
            currency_symbol = options.get("currency_symbol", "₱")
            processed_headers, processed_data = ExportUtility.process_data_with_currency(data, headers, currency_symbol)
            
            # Get save file location
            default_filename = f"{filename_prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filename, _ = QFileDialog.getSaveFileName(
                parent,
                f"Export to Excel",
                default_filename,
                "Excel Files (*.xlsx)"
            )
            
            if not filename:
                return False, "Export cancelled"
            
            # Show progress dialog
            progress = QProgressDialog("Exporting to Excel...", "Cancel", 0, 100, parent)
            progress.setWindowModality(Qt.WindowModal)
            progress.show()
            
            # Get color scheme
            color_scheme = options.get("color_scheme", "blue")
            colors = ExportUtility.COLOR_SCHEMES[color_scheme]
            
            # Convert hex colors to openpyxl colors
            header_fill = PatternFill(start_color=colors["header_fill"], end_color=colors["header_fill"], fill_type="solid")
            alt_row_fill = PatternFill(start_color=colors["alt_row_fill"], end_color=colors["alt_row_fill"], fill_type="solid")
            
            # Create border style
            border = Border(
                left=Side(style='thin', color=colors["border"]),
                right=Side(style='thin', color=colors["border"]),
                top=Side(style='thin', color=colors["border"]),
                bottom=Side(style='thin', color=colors["border"])
            )
            
            # Create workbook and worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name
            
            # Set progress to 10%
            progress.setValue(10)
            
            # Add company information if requested
            row_offset = 0
            if options.get("include_company_info", True):
                company_info = ExportUtility.get_company_info()
                
                # Add company name
                company_cell = ws.cell(row=1, column=1, value=company_info["name"])
                company_cell.font = Font(size=16, bold=True, color=colors["title_font"])
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(processed_headers))
                
                # Try to add logo if it exists
                logo_path = company_info["logo_path"]
                if os.path.exists(logo_path):
                    try:
                        img = XLImage(logo_path)
                        img.width = 100
                        img.height = 100
                        ws.add_image(img, "A2")
                        row_offset += 7  # Space for logo
                    except Exception:
                        # If logo fails, just add more space for company name
                        row_offset += 1
                else:
                    row_offset += 1
            
            # Add date and time if requested
            if options.get("include_datetime", True):
                date_cell = ws.cell(row=row_offset+2, column=1, value=f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
                date_cell.font = Font(italic=True)
                ws.merge_cells(start_row=row_offset+2, start_column=1, end_row=row_offset+2, end_column=len(processed_headers))
                row_offset += 2
            
            # Add report title
            title_cell = ws.cell(row=row_offset+2, column=1, value=sheet_name)
            title_cell.font = Font(size=14, bold=True, color=colors["title_font"])
            ws.merge_cells(start_row=row_offset+2, start_column=1, end_row=row_offset+2, end_column=len(processed_headers))
            row_offset += 3
            
            # Set progress to 20%
            progress.setValue(20)
            
            # Add headers
            header_font = Font(bold=True, color=colors["header_font"], size=options.get("font_size", 10) + 1)
            for col_idx, header in enumerate(processed_headers, 1):
                cell = ws.cell(row=row_offset, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = border
            
            # Set progress to 30%
            progress.setValue(30)
            
            # Add data rows
            row_font = Font(size=options.get("font_size", 10))
            row_alignment = Alignment(vertical='center')
            
            for row_idx, row_data in enumerate(processed_data, 1):
                # Update progress (30% to 80%)
                progress_value = 30 + int(50 * row_idx / len(processed_data))
                progress.setValue(progress_value)
                
                # Check if user cancelled
                if progress.wasCanceled():
                    return False, "Export cancelled by user"
                
                # Add row data
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_offset+row_idx, column=col_idx, value=value)
                    cell.font = row_font
                    cell.alignment = row_alignment
                    cell.border = border
                    
                    # Apply alternate row styling if enabled
                    if options.get("alternate_rows", True) and row_idx % 2 == 0:
                        cell.fill = alt_row_fill
            
            # Set progress to 80%
            progress.setValue(80)
            
            # Include totals if requested
            if options.get("include_totals", False):
                total_row = row_offset + len(processed_data) + 1
                total_label = ws.cell(row=total_row, column=1, value="Total")
                total_label.font = Font(bold=True, size=options.get("font_size", 10))
                total_label.border = border
                
                # Add SUM formulas for numeric columns
                for col_idx, header in enumerate(processed_headers, 1):
                    if col_idx > 1:  # Skip first column
                        # Check if column contains numeric data
                        numeric_column = True
                        for row_idx, row_data in enumerate(processed_data, 1):
                            if row_idx <= 5:  # Check first 5 rows
                                try:
                                    value = row_data[col_idx-1]
                                    # Skip if it's just a currency symbol
                                    if isinstance(value, str) and value.strip() in ["$", "₱", "€", "£"]:
                                        continue
                                    # Try to extract numeric value from currency string (e.g., "₱123.45")
                                    if isinstance(value, str) and any(sym in value for sym in ["$", "₱", "€", "£"]):
                                        numeric_value = re.sub(r'[^\d.]', '', value)
                                        float(numeric_value)
                                    else:
                                        float(value)
                                except (ValueError, TypeError, IndexError):
                                    numeric_column = False
                                    break
                        
                        if numeric_column:
                            cell = ws.cell(row=total_row, column=col_idx)
                            start_row = row_offset + 1
                            end_row = row_offset + len(processed_data)
                            cell.value = f"=SUM({get_column_letter(col_idx)}{start_row}:{get_column_letter(col_idx)}{end_row})"
                            cell.font = Font(bold=True, size=options.get("font_size", 10))
                            cell.border = border
                            cell.fill = header_fill
                            cell.font = header_font
            
            # Auto-fit columns if requested
            if options.get("autofit_columns", True):
                for col_idx in range(1, len(processed_headers) + 1):
                    max_length = 0
                    column_letter = get_column_letter(col_idx)
                    
                    # Check header length
                    header_text = str(processed_headers[col_idx-1])
                    if len(header_text) > max_length:
                        max_length = len(header_text)
                    
                    # Check data rows
                    for row_idx, row_data in enumerate(processed_data):
                        if col_idx <= len(row_data):
                            cell_value = str(row_data[col_idx-1])
                            if len(cell_value) > max_length:
                                max_length = len(cell_value)
                    
                    # Adjust column width with some padding
                    adjusted_width = max_length + 2
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            # Freeze header row if requested
            if options.get("freeze_header", True):
                freeze_cell = f'A{row_offset+1}'
                ws.freeze_panes = ws[freeze_cell]
            
            # Set progress to 90%
            progress.setValue(90)
            
            # Save the workbook
            wb.save(filename)
            
            # Set progress to 100%
            progress.setValue(100)
            
            return True, f"Data exported successfully to {filename}"
            
        except Exception as e:
            return False, f"Failed to export data: {str(e)}"
        finally:
            if 'progress' in locals():
                progress.close()
    
    @staticmethod
    def export_to_pdf(parent, data, headers, filename_prefix, title=None, subtitle=None, options=None):
        """
        Enhanced export to PDF file with styling options
        
        Parameters:
        - parent: Parent widget for file dialog
        - data: List of data rows
        - headers: List of column headers
        - filename_prefix: Prefix for the default filename
        - title: Title of the PDF document (optional)
        - subtitle: Optional subtitle
        - options: Dictionary of export options (optional)
        
        Returns:
        - tuple: (bool: success, str: message)
        """
        try:
            # Show export options dialog if not provided
            if options is None:
                dialog = ExportOptionsDialog(parent, export_type="pdf")
                if dialog.exec_() != QDialog.Accepted:
                    return False, "Export cancelled"
                
                options = dialog.get_options()
                if not title:
                    title = options["title"]
            elif not title:
                title = "Pharmacy Management System Report"
            
            # Process data to replace currency symbols
            currency_symbol = options.get("currency_symbol", "₱")
            processed_headers, processed_data = ExportUtility.process_data_with_currency(data, headers, currency_symbol)
            
            # Get save file location
            default_filename = f"{filename_prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            filename, _ = QFileDialog.getSaveFileName(
                parent,
                f"Export to PDF",
                default_filename,
                "PDF Files (*.pdf)"
            )
            
            if not filename:
                return False, "Export cancelled"
            
            # Show progress dialog
            progress = QProgressDialog("Exporting to PDF...", "Cancel", 0, 100, parent)
            progress.setWindowModality(Qt.WindowModal)
            progress.show()
            
            # Set page size
            paper_size = options.get("paper_size", "Letter")
            orientation = options.get("orientation", "portrait")
            
            if paper_size == "Letter":
                page_size = letter
            elif paper_size == "A4":
                page_size = A4
            elif paper_size == "Legal":
                page_size = (8.5*inch, 14*inch)
            
            if orientation == "landscape":
                page_size = landscape(page_size)
            
            # Get color scheme
            color_scheme = options.get("color_scheme", "blue")
            colors_dict = ExportUtility.COLOR_SCHEMES[color_scheme]
            
            # Convert hex colors to reportlab colors
            header_color = colors.HexColor(f"#{colors_dict['header_fill']}")
            header_font_color = colors.HexColor(f"#{colors_dict['header_font']}")
            alt_row_color = colors.HexColor(f"#{colors_dict['alt_row_fill']}")
            border_color = colors.HexColor(f"#{colors_dict['border']}")
            title_color = colors.HexColor(f"#{colors_dict['title_font']}")
            
            # Create PDF document
            doc = SimpleDocTemplate(filename, pagesize=page_size, leftMargin=0.5*inch, rightMargin=0.5*inch)
            elements = []
            
            # Set progress to 10%
            progress.setValue(10)
            
            # Add styles
            styles = getSampleStyleSheet()
            
            # Create custom styles
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                textColor=title_color,
                spaceAfter=12
            )
            
            subtitle_style = ParagraphStyle(
                'CustomSubtitle',
                parent=styles['Heading2'],
                fontSize=14,
                textColor=title_color,
                spaceAfter=12
            )
            
            normal_style = ParagraphStyle(
                'CustomNormal',
                parent=styles['Normal'],
                fontSize=options.get("font_size", 10),
                spaceAfter=6
            )
            
            # Add company information if requested
            if options.get("include_company_info", True):
                company_info = ExportUtility.get_company_info()
                
                # Try to add logo if it exists
                logo_path = company_info["logo_path"]
                if os.path.exists(logo_path):
                    try:
                        logo = Image(logo_path, width=1*inch, height=1*inch)
                        elements.append(logo)
                        elements.append(Spacer(1, 12))
                    except Exception:
                        pass
                
                # Add company name
                company_name = Paragraph(company_info["name"], title_style)
                elements.append(company_name)
            
            # Set progress to 20%
            progress.setValue(20)
            
            # Add title
            title_para = Paragraph(title, title_style if not options.get("include_company_info", True) else subtitle_style)
            elements.append(title_para)
            
            # Add subtitle if provided
            if subtitle:
                subtitle_para = Paragraph(subtitle, subtitle_style)
                elements.append(subtitle_para)
            
            # Add date if requested
            if options.get("include_datetime", True):
                date_text = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
                date_para = Paragraph(date_text, normal_style)
                elements.append(date_para)
            
            elements.append(Spacer(1, 12))
            
            # Set progress to 30%
            progress.setValue(30)
            
            # Create table with headers
            table_data = [processed_headers]
            
            # Add data rows
            table_data.extend(processed_data)
            
            # Create table
            table = Table(table_data)
            
            # Add table style
            style = [
                ('BACKGROUND', (0, 0), (-1, 0), header_color),
                ('TEXTCOLOR', (0, 0), (-1, 0), header_font_color),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), options.get("font_size", 10) + 2),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 1, border_color),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), options.get("font_size", 10)),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]
            
            # Add alternating row colors if enabled
            if options.get("alternate_rows", True):
                for i in range(1, len(table_data)):
                    if i % 2 == 0:
                        style.append(('BACKGROUND', (0, i), (-1, i), alt_row_color))
            
            # Include totals if requested
            if options.get("include_totals", False):
                # Calculate totals for numeric columns
                totals = ["Total"]
                
                for col_idx in range(1, len(processed_headers)):
                    try:
                        # Identify if this is a numeric column
                        numeric_values = []
                        for row in processed_data:
                            value = row[col_idx]
                            # Skip if it's just a currency symbol
                            if isinstance(value, str) and value.strip() in ["$", "₱", "€", "£"]:
                                continue
                            
                            # Extract numeric value from currency string if needed
                            if isinstance(value, str):
                                if any(sym in value for sym in ["$", "₱", "€", "£"]):
                                    try:
                                        value = float(re.sub(r'[^\d.]', '', value))
                                    except ValueError:
                                        continue
                                elif value.replace('.', '', 1).isdigit():
                                    value = float(value)
                                else:
                                    continue
                            
                            if isinstance(value, (int, float)):
                                numeric_values.append(value)
                        
                        # Calculate total if we have numeric values
                        if numeric_values:
                            total = sum(numeric_values)
                            # Format with appropriate currency symbol if needed
                            if any(isinstance(row[col_idx], str) and any(sym in row[col_idx] for sym in ["$", "₱", "€", "£"]) 
                                  for row in processed_data[:5]):  # Check first 5 rows
                                totals.append(f"{currency_symbol}{total:.2f}")
                            else:
                                totals.append(f"{total:.2f}")
                        else:
                            totals.append("")
                    except (IndexError, ValueError, TypeError):
                        totals.append("")
                
                # Add total row
                table_data.append(totals)
                
                # Add style for total row
                total_row = len(table_data) - 1  # Last row
                style.append(('BACKGROUND', (0, total_row), (-1, total_row), header_color))
                style.append(('TEXTCOLOR', (0, total_row), (-1, total_row), header_font_color))
                style.append(('FONTNAME', (0, total_row), (-1, total_row), 'Helvetica-Bold'))
            
            table.setStyle(TableStyle(style))
            elements.append(table)
            
            # Set progress to 90%
            progress.setValue(90)
            
            # Build PDF
            doc.build(elements)
            
            # Set progress to 100%
            progress.setValue(100)
            
            return True, f"Data exported successfully to {filename}"
            
        except Exception as e:
            return False, f"Failed to export data to PDF: {str(e)}"
        finally:
            if 'progress' in locals():
                progress.close()
    
    @staticmethod
    def export_data(parent, data, headers, filename_prefix, format_type="excel", **kwargs):
        """
        Universal export function that handles both Excel and PDF exports with a consistent interface
        
        Parameters:
        - parent: Parent widget for file dialog
        - data: List of data rows
        - headers: List of column headers
        - filename_prefix: Prefix for the default filename
        - format_type: "excel" or "pdf"
        - **kwargs: Additional options for specific export formats
        
        Returns:
        - bool: True if export was successful, False otherwise
        """
        try:
            # Check if there's data to export
            if not data or not headers:
                QMessageBox.warning(parent, "Export Error", "No data to export.")
                return False
            
            # Determine export function
            if format_type.lower() == "excel":
                success, message = ExportUtility.export_to_excel(
                    parent, data, headers, filename_prefix, 
                    kwargs.get('sheet_name'), kwargs.get('options')
                )
            elif format_type.lower() == "pdf":
                success, message = ExportUtility.export_to_pdf(
                    parent, data, headers, filename_prefix,
                    kwargs.get('title'), kwargs.get('subtitle'), kwargs.get('options')
                )
            else:
                QMessageBox.warning(parent, "Export Error", f"Unsupported export format: {format_type}")
                return False
            
            # Show result message
            if success:
                QMessageBox.information(parent, "Export Successful", message)
            else:
                QMessageBox.critical(parent, "Export Error", message)
            
            return success
            
        except Exception as e:
            QMessageBox.critical(parent, "Export Error", f"Failed to export data: {str(e)}")
            return False